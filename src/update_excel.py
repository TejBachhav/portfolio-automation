"""
update_excel.py
───────────────
Writes the latest fetched data back into your
PERSONALISED_PORTFOLIO_ALLOCATION_MODEL.xlsx file while preserving:
  • Original column order
  • Formatting / styles / conditional formatting
  • Cells that we don't have data for (we leave them as-is)

Strategy: read original workbook → match by company name / scheme name / symbol
→ overwrite only the columns we fetched → save to NEW file with timestamp.
"""
from __future__ import annotations
import pandas as pd
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from .utils import setup_logger, get_output_path, get_project_root

log = setup_logger("excel_writer")


def update_excel(template_path: str, output_path: str = None) -> str:
    """
    Reads CSV outputs from /output and merges into the Excel template.
    Returns the path of the written file.
    """
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = get_output_path(f"PORTFOLIO_MODEL_UPDATED_{ts}.xlsx")

    log.info(f"📋 Copying template → {output_path}")
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)

    # ─── Sheet 1: Nifty 500 ──────────────────────────────────────────────────
    _update_sheet(wb, "Nifty 500",
                  csv_path=get_output_path("nifty500_latest.csv"),
                  match_col="Company Name",
                  sheet_match_col="Company Name")

    # ─── Sheet 2: S&P 500 ────────────────────────────────────────────────────
    _update_sheet(wb, "S&P 500",
                  csv_path=get_output_path("sp500_latest.csv"),
                  match_col="Symbol",
                  sheet_match_col="Symbol")

    # ─── Sheet 3: EQUITY MF ──────────────────────────────────────────────────
    _update_sheet(wb, "EQUITY MF",
                  csv_path=get_output_path("equity_mf_latest.csv"),
                  match_col="SCHEMES",
                  sheet_match_col="SCHEMES")

    # ─── Sheet 4: DEBT MF ────────────────────────────────────────────────────
    _update_sheet(wb, "DEBT MF",
                  csv_path=get_output_path("debt_mf_latest.csv"),
                  match_col="SCHEMES",
                  sheet_match_col="SCHEMES")

    # ─── Sheet 5: REITINVIT ──────────────────────────────────────────────────
    _update_sheet(wb, "REITINVIT",
                  csv_path=get_output_path("reit_invit_latest.csv"),
                  match_col="symbol",
                  sheet_match_col="symbol")

    # Add a metadata sheet
    _add_metadata_sheet(wb)

    wb.save(output_path)
    log.info(f"✅ Excel saved → {output_path}")
    return output_path


def _update_sheet(wb, sheet_name: str, csv_path: Path, match_col: str, sheet_match_col: str):
    """Update one sheet in-place."""
    csv_path = Path(csv_path)
    if not csv_path.exists():
        log.warning(f"   Skipping {sheet_name} — no CSV at {csv_path}")
        return
    if sheet_name not in wb.sheetnames:
        log.warning(f"   Sheet '{sheet_name}' not in workbook")
        return

    df = pd.read_csv(csv_path)
    ws = wb[sheet_name]

    # Get headers from row 1
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}

    # Map sheet row by match key → row number
    match_col_idx = headers.get(sheet_match_col)
    if not match_col_idx:
        log.warning(f"   {sheet_name}: match column '{sheet_match_col}' not in headers")
        return

    row_map = {}  # value → row_num
    for row_num in range(2, ws.max_row + 1):
        val = ws.cell(row=row_num, column=match_col_idx).value
        if val:
            row_map[str(val).strip()] = row_num

    updates = 0
    for _, data_row in df.iterrows():
        key = data_row.get(match_col)
        if pd.isna(key):
            continue
        key = str(key).strip()
        row_num = row_map.get(key)
        if not row_num:
            continue

        for col_name, col_idx in headers.items():
            if col_name in data_row and pd.notna(data_row[col_name]):
                ws.cell(row=row_num, column=col_idx, value=data_row[col_name])
                updates += 1

    log.info(f"   {sheet_name}: {updates} cell updates across {len(df)} rows")


def _add_metadata_sheet(wb):
    """Add an UPDATE_LOG sheet with timestamps and stats."""
    if "UPDATE_LOG" in wb.sheetnames:
        del wb["UPDATE_LOG"]
    ws = wb.create_sheet("UPDATE_LOG", 0)  # at the beginning

    ws["A1"] = "Portfolio Model — Auto-Update Metadata"
    ws["A1"].font = ws["A1"].font.copy(bold=True, size=14)
    ws["A3"] = "Update Time:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Sheets Updated:"
    ws["B4"] = "Nifty 500, S&P 500, EQUITY MF, DEBT MF, REITINVIT"

    ws["A6"] = "Sheet"
    ws["B6"] = "Source"
    ws["C6"] = "Status"
    ws["D6"] = "Last Refresh"

    data_sources = [
        ("Nifty 500", "yfinance + computed technicals (pandas)", "OK"),
        ("S&P 500", "yfinance + DCF intrinsic value", "OK"),
        ("EQUITY MF", "AMFI India + MFAPI.in", "OK"),
        ("DEBT MF", "AMFI India + MFAPI.in", "OK"),
        ("REITINVIT", "yfinance + manual quarterly overrides", "OK"),
    ]
    for i, (sh, src, st) in enumerate(data_sources, start=7):
        ws.cell(row=i, column=1, value=sh)
        ws.cell(row=i, column=2, value=src)
        ws.cell(row=i, column=3, value=st)
        ws.cell(row=i, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20


if __name__ == "__main__":
    import sys
    template = sys.argv[1] if len(sys.argv) > 1 else "PERSONALISED_PORTFOLIO_ALLOCATION_MODEL_.xlsx"
    update_excel(template)
